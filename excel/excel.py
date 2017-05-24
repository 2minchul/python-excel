#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
import xlrd
import os

try:
    from .exceptions import *
except:
    from exceptions import * 


def Excel(filename = '', mode = 'r'):
    if filename:
        ext = os.path.splitext(filename)[-1]
        if ext == '.xls':
            return XlsController(filename, mode)
        elif ext != '.xlsx':
            raise ReadError(filename)

    return XlsxController(filename, mode)

class XlsxCell:
    def __init__(self, cell):
        self.cell = cell
        self.col_idx = cell.col_idx
        self.value = cell._value
    
    def get_value(self):
        return self.cell._value
    
    def set_text(self, value):
        self.cell.number_format = '@'
        self.cell.value = value
    
    def set_value(self, value):
        self.cell.value = value
    
    def __str__(self):
        return self.cell.__str__()
    
    def __repr__(self):
        return self.cell.__repr__()

class XlsCell:
    def __init__(self, cell, row_n, col_n, temp):
        self.cell = cell
        self.col_idx = col_n
        self.row_idx = row_n
        self.set_text = self.set_value
        self.memory = temp
        self.value = self.get_value()
    
    def get_value(self):
        buff = self.memory.get((self.row_idx, self.col_idx))
        if buff:
            return buff
        return self.cell.value
    
    def set_value(self, value):
        self.memory[(self.row_idx, self.col_idx)] = value
    
    def __str__(self):
        return self.cell.__str__()
    
    def __repr__(self):
        return self.cell.__repr__()
    


class XlsxController:
    def __init__(self, filename = '', mode = 'r'):
        self.filename = filename
        self.mode = mode
        
        if filename and mode == 'r':
            if not self.load(filename):
                raise ReadError(filename)
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
        
        self.max_row = self.height = self.ws.max_row
        self.max_col = self.width = self.ws.max_column
        
        self.cell = self.get_cell
        

    def load(self, filename):
        try:
            self.wb = openpyxl.load_workbook(filename)
            self.ws = self.wb.active
        except:
            return False
        else:
            return True
    
    def save(self, filename):
        print('write Excel file: %s' % (os.path.basename(filename)))
        self.wb.save(filename)
    
    def get_cell(self, row_n, col_n):
        return XlsxCell(self.ws.cell(row = row_n + 1, column = col_n + 1))
        
    def get_row(self, row_n = 0):
        return tuple(XlsxCell(cell) for cell in self.ws[str(row_n+1)])
    
    def get_col(self, col_n = 0):
        return tuple(XlsxCell(cell) for cell in self.ws[self._num2col(col_n + 1)])
    
    def set_freeze_panes(self, index):
        self.ws.freeze_panes = index
    
    def auto_resize_col(self):
        def as_text(value): return str(value) if value is not None else ""
        
        for column_cells in self.ws.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            self.ws.column_dimensions[column_cells[0].column].width = length + 2        
        
    
    def _num2col(self, col):
        return openpyxl.cell.cell.get_column_letter(col)
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.mode == 'w' and self.filename:
            self.save(self.filename)
        
        del self.ws
        self.wb.close()


class XlsController:
    def __init__(self, filename = '', mode = 'r'):
        self.filename = filename
        self.mode = mode
        self.temp_memory = {}
        
        if filename and mode == 'r':
            if not self.load(filename):
                raise ReadError(filename)
        else:
            raise ReadError(filename)
            
        
        
        self.max_row = self.height = self.ws.nrows
        self.max_col = self.width = self.ws.ncols
        self.cell = self.get_cell

    
    def load(self, filename):
        try:
            self.wb = xlrd.open_workbook(filename)
            self.ws = self.wb.sheet_by_index(0)
        except:
            return False
        else:
            return True
    
    def save(self, filename):
        # Todo
        #self.wb.save(filename)
        self.save_as_xlsx(filename)
    
    def save_as_xlsx(self, filename):
        filename = os.path.splitext(filename)[0] + '.xlsx'
        with Excel(filename, 'w') as sheet:
            for row_n in range(self.max_row):
                for col_n in range(self.max_col):
                    sheet.cell(row_n, col_n).set_text(self.cell(row_n, col_n).get_value())
            sheet.auto_resize_col()
    
    def get_cell(self, row_n, col_n):
        return XlsCell(self.ws.cell(row_n, col_n), row_n, col_n, self.temp_memory)
    
    def get_row(self, row_n = 0):
        return tuple(XlsCell(self.ws.cell(row_n, col_n), row_n, col_n, self.temp_memory) for col_n in range(self.max_col))
    
    def get_col(self, col_n = 0):
        return tuple(XlsCell(self.ws.cell(row_n, col_n), row_n, col_n, self.temp_memory) for row_n in range(self.max_row))
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.mode == 'w' and self.filename:
            self.save(self.filename)
        
        del self.ws
        del self.wb

if __name__ == '__main__':
    s = Excel('D:/form.xlsx')
    for k in s.get_row(0):
        print(k.cell.number_format)
    
    